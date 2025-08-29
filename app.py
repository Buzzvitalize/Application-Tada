from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    send_file,
    send_from_directory,
    flash,
    session,
    jsonify,
    g,
    current_app,
    Response,
    stream_with_context,
)
from flask_migrate import Migrate, upgrade
import logging
from logging.handlers import RotatingFileHandler
import smtplib
from email.mime.text import MIMEText
from flask_wtf import CSRFProtect
from models import (
    db,
    migrate,
    Client,
    Product,
    Quotation,
    QuotationItem,
    Order,
    OrderItem,
    Invoice,
    InvoiceItem,
    Payment,
    InventoryMovement,
    Warehouse,
    ProductStock,
    CompanyInfo,
    User,
    AccountRequest,
    ExportLog,
    NcfLog,
    Notification,
    dom_now,
)
from io import BytesIO, StringIO
import csv
try:
    from openpyxl import Workbook
except ModuleNotFoundError:  # pragma: no cover
    Workbook = None
from datetime import datetime, timedelta
from sqlalchemy import func, inspect, or_
from sqlalchemy.exc import NoSuchTableError
from sqlalchemy.orm import load_only, joinedload
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
import os
import re
import json
from ai import recommend_products
from weasy_pdf import generate_pdf
from account_pdf import generate_account_statement_pdf
from functools import wraps
from auth import auth_bp, generate_reset_token
from forms import AccountRequestForm
from config import DevelopmentConfig
try:
    from dotenv import load_dotenv
except ModuleNotFoundError:  # pragma: no cover
    def load_dotenv():
        pass
try:
    from rq import Queue
    from redis import Redis
except Exception:  # pragma: no cover
    Queue = None
    Redis = None
import threading

load_dotenv()
# Load RNC data for company name lookup
RNC_DATA = {}
DATA_PATH = os.path.join(os.path.dirname(__file__), 'data', 'DGII_RNC.TXT')
if os.path.exists(DATA_PATH):
    with open(DATA_PATH, encoding='utf-8') as f:
        for row in f:
            parts = row.strip().split('|')
            if len(parts) >= 2:
                rnc = re.sub(r'\D', '', parts[0])
                name = parts[1].strip()
                if rnc:
                    RNC_DATA[rnc] = name

app = Flask(__name__)
app.config.from_object(DevelopmentConfig)

if not os.path.exists('logs'):
    os.makedirs('logs')
file_handler = RotatingFileHandler('logs/app.log', maxBytes=10240, backupCount=10)
file_handler.setFormatter(logging.Formatter('%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'))
file_handler.setLevel(logging.INFO)
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)
app.logger.info('Tiendix startup')

SMTP_HOST = os.getenv('SMTP_HOST')
SMTP_PORT = int(os.getenv('SMTP_PORT', 587))
SMTP_USER = os.getenv('SMTP_USER')
SMTP_PASS = os.getenv('SMTP_PASS')
SMTP_FROM = os.getenv('SMTP_FROM', SMTP_USER)


def send_email(to, subject, html):
    if not SMTP_HOST or not SMTP_FROM:
        app.logger.warning('Email settings missing; skipping send to %s', to)
        return
    msg = MIMEText(html, 'html')
    msg['Subject'] = subject
    msg['From'] = SMTP_FROM
    msg['To'] = to
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            if SMTP_USER and SMTP_PASS:
                s.starttls()
                s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(SMTP_FROM, [to], msg.as_string())
    except Exception as e:  # pragma: no cover
        app.logger.error('Email send failed: %s', e)


def _fmt_money(value):
    return f"RD${value:,.2f}"


app.jinja_env.filters['money'] = _fmt_money

db.init_app(app)
migrate.init_app(app, db)
csrf = CSRFProtect(app)
app.register_blueprint(auth_bp)

# The database schema is managed via Flask-Migrate.  Tables should be
# created with ``flask db upgrade`` instead of ``db.create_all`` to avoid
# diverging from migrations.

if Queue and Redis:
    redis_conn = Redis.from_url(os.getenv('REDIS_URL', 'redis://localhost:6379'))
    export_queue = Queue('exports', connection=redis_conn)
else:  # pragma: no cover
    export_queue = None


def enqueue_export(fn, *args):
    """Enqueue an export job using RQ if available or fallback to threading."""
    app_obj = current_app._get_current_object()
    if export_queue:
        return export_queue.enqueue(fn, app_obj, *args)
    t = threading.Thread(target=fn, args=(app_obj, *args), daemon=True)
    t.start()
    return t


def log_export(user, formato, tipo, filtros, status, message='', file_path=None):
    with app.app_context():
        entry = ExportLog(
            user=user,
            company_id=current_company_id(),
            formato=formato,
            tipo=tipo,
            filtros=json.dumps(filtros),
            status=status,
            message=message,
            file_path=file_path,
        )
        db.session.add(entry)
        db.session.commit()
        return entry.id


def _export_job(app_obj, company_id, user, start, end, estado, categoria, formato, tipo, entry_id):  # pragma: no cover - background
    """Background task that builds the export file for the given company."""
    with app_obj.app_context():
        filtros = {
            'fecha_inicio': start.strftime('%Y-%m-%d') if start else '',
            'fecha_fin': end.strftime('%Y-%m-%d') if end else '',
            'estado': estado or '',
            'categoria': categoria or '',
        }
        try:
            if os.path.isfile('maint'):
                os.remove('maint')
            os.makedirs('maint', exist_ok=True)
            q = Invoice.query.filter_by(company_id=company_id)
            if start:
                q = q.filter(Invoice.date >= start)
            if end:
                q = q.filter(Invoice.date <= end)
            if estado:
                q = q.filter(Invoice.status == estado)
            if categoria:
                q = q.join(Invoice.items).filter(InvoiceItem.category == categoria)
            path = os.path.join('maint', f'export_{entry_id}.{formato}')
            if formato == 'csv':
                with open(path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    if tipo == 'resumen':
                        writer.writerow(['Categoría', 'Cantidad', 'Total'])
                        summary = (
                            company_query(InvoiceItem)
                            .join(Invoice)
                            .with_entities(
                                InvoiceItem.category,
                                func.count(InvoiceItem.id),
                                func.sum(InvoiceItem.unit_price * InvoiceItem.quantity - InvoiceItem.discount),
                            )
                            .group_by(InvoiceItem.category)
                        )
                        if start:
                            summary = summary.filter(Invoice.date >= start)
                        if end:
                            summary = summary.filter(Invoice.date <= end)
                        if estado:
                            summary = summary.filter(Invoice.status == estado)
                        if categoria:
                            summary = summary.filter(InvoiceItem.category == categoria)
                        for cat, cnt, tot in summary:
                            writer.writerow([cat or 'Sin categoría', cnt, f"{tot or 0:.2f}"])
                    else:
                        writer.writerow(['Cliente', 'Fecha', 'Estado', 'Total'])
                        stream_q = q.options(
                            joinedload(Invoice.client),
                            load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
                        ).yield_per(100)
                        for inv in stream_q:
                            writer.writerow([
                                inv.client.name if inv.client else '',
                                inv.date.strftime('%Y-%m-%d'),
                                inv.status or '',
                                f"{inv.total:.2f}",
                            ])
            elif formato == 'xlsx' and Workbook is not None:
                wb = Workbook()
                ws = wb.active
                if tipo == 'resumen':
                    ws.append(['Categoría', 'Cantidad', 'Total'])
                    summary = (
                        company_query(InvoiceItem)
                        .join(Invoice)
                        .with_entities(
                            InvoiceItem.category,
                            func.count(InvoiceItem.id),
                            func.sum(InvoiceItem.unit_price * InvoiceItem.quantity - InvoiceItem.discount),
                        )
                        .group_by(InvoiceItem.category)
                    )
                    if start:
                        summary = summary.filter(Invoice.date >= start)
                    if end:
                        summary = summary.filter(Invoice.date <= end)
                    if estado:
                        summary = summary.filter(Invoice.status == estado)
                    if categoria:
                        summary = summary.filter(InvoiceItem.category == categoria)
                    for cat, cnt, tot in summary:
                        ws.append([cat or 'Sin categoría', cnt, float(tot or 0)])
                else:
                    ws.append(['Cliente', 'Fecha', 'Estado', 'Total'])
                    stream_q = q.options(
                        joinedload(Invoice.client),
                        load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
                    ).yield_per(100)
                    for inv in stream_q:
                        ws.append([
                            inv.client.name if inv.client else '',
                            inv.date.strftime('%Y-%m-%d'),
                            inv.status or '',
                            float(inv.total),
                        ])
                wb.save(path)
            entry = ExportLog.query.get(entry_id)
            entry.status = 'success'
            entry.file_path = path
            db.session.commit()
        except Exception as exc:  # pragma: no cover - hard to simulate failures
            entry = ExportLog.query.get(entry_id)
            entry.status = 'fail'
            entry.message = str(exc)
            db.session.commit()
def _migrate_legacy_schema():
    """Add missing columns to older SQLite databases.

    Early versions of the project lacked fields such as ``Product.category``.
    Users with an old ``database.sqlite`` would see errors like
    "no such column: product.category" when creating cotizaciones.  This helper
    checks for expected columns and adds them on the fly so the application can
    continue running without manual intervention.
    """
    inspector = inspect(db.engine)
    statements = []

    if inspector.has_table('product'):
        try:
            product_cols = {c['name'] for c in inspector.get_columns('product')}
        except NoSuchTableError:  # pragma: no cover - sqlite reflection race
            product_cols = set()
        if 'category' not in product_cols:
            statements.append("ALTER TABLE product ADD COLUMN category VARCHAR(50)")
        if 'unit' not in product_cols:
            statements.append("ALTER TABLE product ADD COLUMN unit VARCHAR(20) DEFAULT 'Unidad'")
        if 'has_itbis' not in product_cols:
            statements.append("ALTER TABLE product ADD COLUMN has_itbis BOOLEAN DEFAULT 1")

    if inspector.has_table('user'):
        try:
            user_cols = {c['name'] for c in inspector.get_columns('user')}
        except NoSuchTableError:  # pragma: no cover - sqlite reflection race
            user_cols = set()
        if 'email' not in user_cols:
            statements.append("ALTER TABLE user ADD COLUMN email VARCHAR(120)")
        if 'first_name' not in user_cols:
            statements.append("ALTER TABLE user ADD COLUMN first_name VARCHAR(120) DEFAULT ''")
        if 'last_name' not in user_cols:
            statements.append("ALTER TABLE user ADD COLUMN last_name VARCHAR(120) DEFAULT ''")

    if inspector.has_table('inventory_movement'):
        try:
            im_cols = {c['name'] for c in inspector.get_columns('inventory_movement')}
        except NoSuchTableError:  # pragma: no cover - sqlite reflection race
            im_cols = set()
        if 'executed_by' not in im_cols:
            statements.append(
                "ALTER TABLE inventory_movement ADD COLUMN executed_by INTEGER REFERENCES user(id)"
            )

    for stmt in statements:
        db.session.execute(db.text(stmt))
    if statements:
        db.session.commit()


def ensure_admin():  # pragma: no cover - optional helper for deployments
    with app.app_context():
        try:  # Apply any pending migrations for safety
            upgrade()
        except Exception:  # pragma: no cover - fallback when migrations misconfigured
            db.create_all()
        inspector = inspect(db.engine)
        _migrate_legacy_schema()
        if inspector.has_table('user') and not User.query.filter_by(username='admin').first():
            admin = User(username='admin', role='admin', first_name='Admin', last_name='')
            admin.set_password(os.environ.get('ADMIN_PASSWORD', '363636'))
            db.session.add(admin)
            db.session.commit()
        db.session.remove()

# Utility constants
ITBIS_RATE = 0.18
UNITS = ('Unidad', 'Metro', 'Onza', 'Libra', 'Kilogramo', 'Litro')
CATEGORIES = (
    'Alimentos y Bebidas',
    'Productos Industriales / Materiales',
    'Minerales',
    'Salud y Cuidado Personal',
    'Electrónica y Tecnología',
    'Hogar y Construcción',
    'Energía Renovable',
    'Otros',
)
INVOICE_STATUSES = ('Pendiente', 'Pagada')
MAX_EXPORT_ROWS = 50000


def current_company_id():
    return session.get('company_id')


def notify(message):
    if current_company_id():
        db.session.add(Notification(company_id=current_company_id(), message=message))
        db.session.commit()


def company_query(model):
    cid = current_company_id()
    if session.get('role') == 'admin' and cid is None:
        return model.query
    return model.query.filter_by(company_id=cid)


def company_get(model, object_id):
    return company_query(model).filter_by(id=object_id).first_or_404()

def _to_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def generate_reference(name: str) -> str:
    """Generate a unique reference based on product name."""
    prefix = ''.join(ch for ch in (name or '').upper() if ch.isalnum())[:3]
    if not prefix:
        prefix = 'REF'
    existing = company_query(Product).filter(Product.reference.like(f"{prefix}%")).all()
    numbers = []
    for p in existing:
        if p.reference and p.reference.startswith(prefix):
            suf = p.reference[len(prefix):]
            if suf.isdigit():
                numbers.append(int(suf))
    next_no = (max(numbers) + 1) if numbers else 1
    return f"{prefix}{next_no:03d}"


def _parse_report_params(fecha_inicio, fecha_fin, estado, categoria):
    """Validate and normalize report filter parameters."""
    start = end = None
    if fecha_inicio:
        try:
            start = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        except ValueError:
            start = None
    if fecha_fin:
        try:
            end = datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            end = None
    if start and end and start > end:
        start = end = None
    if estado not in INVOICE_STATUSES:
        estado = None
    if categoria not in CATEGORIES:
        categoria = None
    return start, end, estado, categoria


@app.template_filter('phone')
def fmt_phone(value):
    digits = re.sub(r'\D', '', value or '')
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return value or ''


@app.template_filter('id_doc')
def fmt_id(value):
    digits = re.sub(r'\D', '', value or '')
    if len(digits) == 9:
        return f"{digits[:3]}-{digits[3:8]}-{digits[8:]}"
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:10]}-{digits[10:]}"
    return value or ''


def calculate_totals(items):
    subtotal = 0
    itbis = 0
    for item in items:
        line = (item['unit_price'] * item['quantity']) - item['discount']
        subtotal += line
        if item.get('has_itbis'):
            itbis += line * ITBIS_RATE
    return subtotal, itbis, subtotal + itbis


def build_items(product_ids, quantities, discounts):
    # Convert the list of product ids to integers, ignoring any non-numeric
    # values that may come from malformed form submissions. Previously a
    # stray string (e.g. the product name) would raise ``ValueError`` and the
    # quotation silently failed to save.  Now we skip those entries so the
    # view can provide proper feedback to the user.
    ids: list[int] = []
    for pid in product_ids:
        try:
            if pid:
                ids.append(int(pid))
        except (TypeError, ValueError):
            continue
    products = (
        company_query(Product)
        .options(load_only(
            Product.id,
            Product.code,
            Product.reference,
            Product.name,
            Product.unit,
            Product.price,
            Product.category,
            Product.has_itbis,
        ))
        .filter(Product.id.in_(ids))
        .all()
    )
    prod_map = {str(p.id): p for p in products}
    items = []
    for pid, q, d in zip(product_ids, quantities, discounts):
        product = prod_map.get(pid)
        if not product:
            continue
        qty = _to_int(q)
        percent = _to_float(d)
        discount_amount = product.price * qty * (percent / 100)
        items.append({
            'code': product.code,
            'reference': product.reference,
            'product_name': product.name,
            'unit': product.unit,
            'unit_price': product.price,
            'quantity': qty,
            'discount': discount_amount,
            'category': product.category,
            'has_itbis': product.has_itbis,
            'company_id': current_company_id(),
        })
    return items


@app.route('/api/rnc/<rnc>')
def rnc_lookup(rnc):
    clean = rnc.replace('-', '')
    name = RNC_DATA.get(clean)
    if not name:
        client = Client.query.filter(func.replace(Client.identifier, '-', '') == clean).first()
        name = client.name if client else ''
    return jsonify({'name': name})


@app.before_request
def load_company():
    cid = current_company_id()
    g.company = CompanyInfo.query.get(cid) if cid else None


@app.context_processor
def inject_company():
    notif_count = 0
    try:
        if 'user_id' in session and current_company_id():
            low_stock = (
                company_query(ProductStock)
                .filter(ProductStock.stock <= ProductStock.min_stock, ProductStock.min_stock > 0)
                .all()
            )
            for ps in low_stock:
                msg = f"Stock bajo: {ps.product.name}"
                if not Notification.query.filter_by(company_id=current_company_id(), message=msg).first():
                    db.session.add(Notification(company_id=current_company_id(), message=msg))
            if low_stock:
                db.session.commit()
            notif_count = Notification.query.filter_by(company_id=current_company_id(), is_read=False).count()
    except Exception:
        pass
    return {'company': getattr(g, 'company', None), 'notification_count': notif_count}


def get_company_info():
    c = CompanyInfo.query.get(current_company_id())
    if not c:
        return {}
    return {
        'name': c.name,
        'address': f"{c.street}, {c.sector}, {c.province}",
        'rnc': c.rnc,
        'phone': c.phone,
        'website': c.website,
        'logo': os.path.join(app.static_folder, c.logo) if c.logo else None,
        'ncf_final': c.ncf_final,
        'ncf_fiscal': c.ncf_fiscal,
    }
# Routes
@app.before_request
def require_login():
    allowed = {
        'auth.login',
        'static',
        'request_account',
        'auth.logout',
        'auth.reset_request',
        'auth.reset_password',
        'terminos',
    }
    if request.endpoint not in allowed and 'user_id' not in session:
        return redirect(url_for('auth.login'))
    admin_extra = {'admin_companies', 'select_company', 'clear_company',
                   'admin_requests', 'approve_request', 'reject_request'}
    if session.get('role') == 'admin' and not session.get('company_id') \
            and request.endpoint not in allowed.union(admin_extra):
        return redirect(url_for('admin_companies'))


def admin_only(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Acceso restringido')
            return redirect(url_for('list_quotations'))
        return f(*args, **kwargs)
    return wrapper


def manager_only(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get('role') not in ('admin', 'manager'):
            flash('Acceso restringido')
            return redirect(url_for('list_quotations'))
        return f(*args, **kwargs)
    return wrapper

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    return redirect(url_for('list_quotations'))


@app.route('/solicitar-cuenta', methods=['GET', 'POST'])
def request_account():
    form = AccountRequestForm()
    if form.validate_on_submit():
        if not form.accepted_terms.data:
            flash(
                'Debe aceptar los Términos y Condiciones para crear una cuenta en Tiendix.',
                'request',
            )
            return redirect(url_for('request_account'))
        if request.form.get('password') != request.form.get('confirm_password'):
            flash('Las contraseñas no coinciden', 'request')
            return redirect(url_for('request_account'))
        account_type = request.form['account_type']
        identifier = request.form.get('identifier')
        if not identifier:
            flash('Debe ingresar RNC o Cédula', 'request')
            return redirect(url_for('request_account'))
        req = AccountRequest(
            account_type=account_type,
            first_name=request.form['first_name'],
            last_name=request.form['last_name'],
            company=request.form['company'],
            identifier=identifier,
            phone=request.form['phone'],
            email=request.form['email'],
            address=request.form.get('address'),
            website=request.form.get('website'),
            username=request.form['username'],
            password=generate_password_hash(request.form['password']),
            accepted_terms=True,
            accepted_terms_at=dom_now(),
            accepted_terms_ip=request.remote_addr,
            accepted_terms_user_agent=request.headers.get('User-Agent', ''),
        )
        db.session.add(req)
        db.session.commit()
        flash('Solicitud enviada, espere aprobación', 'login')
        return redirect(url_for('auth.login'))
    elif request.method == 'POST':
        flash(
            'Debe aceptar los Términos y Condiciones para crear una cuenta en Tiendix.',
            'request',
        )
        return redirect(url_for('request_account'))
    return render_template('solicitar_cuenta.html', form=form)


@app.route('/terminos')
def terminos():
    return render_template('terminos.html')


@app.route('/admin/solicitudes')
@admin_only
def admin_requests():
    requests = AccountRequest.query.all()
    return render_template('admin_solicitudes.html', requests=requests)


@app.route('/admin/companies')
@admin_only
def admin_companies():
    companies = CompanyInfo.query.all()
    return render_template('admin_companies.html', companies=companies)


@app.route('/admin/companies/select/<int:company_id>')
@admin_only
def select_company(company_id):
    session['company_id'] = company_id
    return redirect(url_for('list_quotations'))


@app.route('/admin/companies/clear')
@admin_only
def clear_company():
    session.pop('company_id', None)
    return redirect(url_for('admin_companies'))


@app.route('/admin/solicitudes/<int:req_id>/aprobar', methods=['POST'])
@admin_only
def approve_request(req_id):
    req = AccountRequest.query.get_or_404(req_id)
    role = request.form.get('role', 'company')
    username = req.username
    password = req.password
    email = req.email
    company = CompanyInfo(
        name=req.company,
        street=req.address or '',
        sector='',
        province='',
        phone=req.phone,
        rnc=req.identifier or '',
        website=req.website,
        logo='',
    )
    db.session.add(company)
    db.session.flush()
    user = User(username=username, first_name=req.first_name, last_name=req.last_name, role=role, company_id=company.id)
    # ``req.password`` ya contiene el hash generado al recibir la solicitud.
    user.password = password
    db.session.add(user)
    db.session.delete(req)
    db.session.commit()

    # Envío de enlace temporal para establecer o restablecer contraseña
    token = generate_reset_token(user)
    html = render_template(
        'emails/account_approved.html',
        username=username,
        company=company.name,
        login_url=url_for('auth.login', _external=True),
        reset_url=url_for('auth.reset_password', token=token, _external=True),
    )
    send_email(email, 'Tu cuenta ha sido aprobada', html)
    flash('Cuenta aprobada')
    return redirect(url_for('admin_requests'))


@app.route('/admin/solicitudes/<int:req_id>/rechazar', methods=['POST'])
@admin_only
def reject_request(req_id):
    req = AccountRequest.query.get_or_404(req_id)
    db.session.delete(req)
    db.session.commit()
    flash('Solicitud rechazada')
    return redirect(url_for('admin_requests'))


# --- CPanel ---


@app.route('/cpaneltx')
@admin_only
def cpanel_home():
    return render_template('cpaneltx.html')


@app.route('/cpaneltx/users')
@admin_only
def cpanel_users():
    q = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)
    query = User.query
    if q:
        query = query.filter(User.username.ilike(f'%{q}%'))
    users = query.order_by(User.id).paginate(page=page, per_page=10, error_out=False)
    return render_template('cpanel_users.html', users=users, q=q)


@app.post('/cpaneltx/users/<int:user_id>/update')
@admin_only
def cpanel_user_update(user_id):
    user = User.query.get_or_404(user_id)
    email = request.form.get('email')
    password = request.form.get('password')
    if email:
        user.email = email
    if password:
        user.set_password(password)
    db.session.commit()
    flash('Usuario actualizado')
    return redirect(url_for('cpanel_users'))


@app.post('/cpaneltx/users/<int:user_id>/role')
@admin_only
def cpanel_user_role(user_id):
    user = User.query.get_or_404(user_id)
    role = request.form.get('role')
    if role in ('admin', 'manager', 'company'):
        user.role = role
        db.session.commit()
        flash('Rol actualizado')
    return redirect(url_for('cpanel_users'))


@app.post('/cpaneltx/users/<int:user_id>/delete')
@admin_only
def cpanel_user_delete(user_id):
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash('Usuario eliminado')
    return redirect(url_for('cpanel_users'))


@app.route('/cpaneltx/companies')
@admin_only
def cpanel_companies():
    q = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)
    query = CompanyInfo.query
    if q:
        query = query.filter(CompanyInfo.name.ilike(f'%{q}%'))
    companies = query.order_by(CompanyInfo.id).paginate(page=page, per_page=10, error_out=False)
    return render_template('cpanel_companies.html', companies=companies, q=q)


@app.post('/cpaneltx/companies/<int:cid>/delete')
@admin_only
def cpanel_company_delete(cid):
    company = CompanyInfo.query.get_or_404(cid)
    db.session.delete(company)
    db.session.commit()
    flash('Empresa eliminada')
    return redirect(url_for('cpanel_companies'))


@app.route('/cpaneltx/orders')
@admin_only
def cpanel_orders():
    orders = Order.query.options(joinedload(Order.client)).all()
    return render_template('cpanel_orders.html', orders=orders)


@app.post('/cpaneltx/orders/<int:oid>/delete')
@admin_only
def cpanel_order_delete(oid):
    order = Order.query.get_or_404(oid)
    db.session.delete(order)
    db.session.commit()
    flash('Pedido eliminado')
    return redirect(url_for('cpanel_orders'))


@app.route('/cpaneltx/invoices')
@admin_only
def cpanel_invoices():
    invoices = Invoice.query.options(joinedload(Invoice.client)).all()
    return render_template('cpanel_invoices.html', invoices=invoices)


@app.post('/cpaneltx/invoices/<int:iid>/delete')
@admin_only
def cpanel_invoice_delete(iid):
    inv = Invoice.query.get_or_404(iid)
    db.session.delete(inv)
    db.session.commit()
    flash('Factura eliminada')
    return redirect(url_for('cpanel_invoices'))

# Clients CRUD
@app.route('/clientes', methods=['GET', 'POST'])
def clients():
    if request.method == 'POST':
        is_final = request.form.get('type') == 'final'
        identifier = request.form.get('identifier') if not is_final else request.form.get('identifier') or None
        last_name = request.form.get('last_name') if is_final else None
        if not is_final and not identifier:
            flash('El RNC es obligatorio para empresas')
            return redirect(url_for('clients'))
        if identifier:
            exists = company_query(Client).filter(Client.identifier == identifier).first()
            if exists:
                flash('Ya existe un cliente con ese RNC/Cédula')
                return redirect(url_for('clients'))
        email = request.form.get('email')
        if email:
            exists = company_query(Client).filter(Client.email == email).first()
            if exists:
                flash('Ya existe un cliente con ese correo electrónico')
                return redirect(url_for('clients'))
        client = Client(
            name=request.form['name'],
            last_name=last_name,
            identifier=identifier,
            phone=request.form.get('phone'),
            email=request.form.get('email'),
            street=request.form.get('street'),
            sector=request.form.get('sector'),
            province=request.form.get('province'),
            is_final_consumer=is_final,
            company_id=current_company_id()
        )
        db.session.add(client)
        db.session.commit()
        flash('Cliente agregado')
        notify('Cliente agregado')
        return redirect(url_for('clients'))
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    query = company_query(Client)
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Client.name.ilike(like),
                Client.last_name.ilike(like),
                Client.identifier.ilike(like),
                Client.email.ilike(like),
            )
        )
    clients = query.order_by(Client.id).paginate(page=page, per_page=25, error_out=False)
    return render_template('clientes.html', clients=clients, q=q)

@app.route('/clientes/delete/<int:client_id>', methods=['POST'])
def delete_client(client_id):
    client = company_get(Client, client_id)
    db.session.delete(client)
    db.session.commit()
    flash('Cliente eliminado')
    return redirect(url_for('clients'))

@app.route('/clientes/edit/<int:client_id>', methods=['GET', 'POST'])
def edit_client(client_id):
    client = company_get(Client, client_id)
    if request.method == 'POST':
        is_final = request.form.get('type') == 'final'
        identifier = request.form.get('identifier') if not is_final else request.form.get('identifier') or None
        last_name = request.form.get('last_name') if is_final else None
        if not is_final and not identifier:
            flash('El RNC es obligatorio para empresas')
            return redirect(url_for('edit_client', client_id=client.id))
        if identifier:
            exists = company_query(Client).filter(
                Client.identifier == identifier, Client.id != client.id
            ).first()
            if exists:
                flash('Ya existe un cliente con ese RNC/Cédula')
                return redirect(url_for('edit_client', client_id=client.id))
        email = request.form.get('email')
        if email:
            exists = company_query(Client).filter(
                Client.email == email, Client.id != client.id
            ).first()
            if exists:
                flash('Ya existe un cliente con ese correo electrónico')
                return redirect(url_for('edit_client', client_id=client.id))
        client.name = request.form['name']
        client.last_name = last_name
        client.identifier = identifier
        client.phone = request.form.get('phone')
        client.email = request.form.get('email')
        client.street = request.form.get('street')
        client.sector = request.form.get('sector')
        client.province = request.form.get('province')
        client.is_final_consumer = is_final
        db.session.commit()
        flash('Cliente actualizado')
        return redirect(url_for('clients'))
    return render_template('cliente_form.html', client=client)

@app.post('/api/clients')
def api_create_client():
    data = request.get_json() or {}
    if not data.get('name'):
        return {'error': 'El nombre es obligatorio'}, 400
    is_final = data.get('type') == 'final'
    identifier = data.get('identifier') if not is_final else data.get('identifier') or None
    last_name = data.get('last_name') if is_final else None
    if not is_final and not identifier:
        return {'error': 'El RNC es obligatorio para empresas'}, 400
    if identifier:
        exists = company_query(Client).filter(Client.identifier == identifier).first()
        if exists:
            return {'error': 'Identifier already exists'}, 400
    email = data.get('email')
    if email:
        exists = company_query(Client).filter(Client.email == email).first()
        if exists:
            return {'error': 'Email already exists'}, 400
    client = Client(
        name=data.get('name'),
        last_name=last_name,
        identifier=identifier,
        phone=data.get('phone'),
        email=data.get('email'),
        street=data.get('street'),
        sector=data.get('sector'),
        province=data.get('province'),
        is_final_consumer=is_final,
        company_id=current_company_id()
    )
    db.session.add(client)
    db.session.commit()
    return {'id': client.id, 'name': client.name, 'identifier': client.identifier}


@app.get('/api/reference')
def api_reference():
    name = request.args.get('name', '')
    return {'reference': generate_reference(name)}

# Products CRUD
@app.route('/productos', methods=['GET', 'POST'])
def products():
    if request.method == 'POST':
        reference = request.form.get('reference') or generate_reference(request.form['name'])
        product = Product(
            code=request.form['code'],
            reference=reference,
            name=request.form['name'],
            unit=request.form['unit'],
            price=_to_float(request.form['price']),
            category=request.form.get('category'),
            has_itbis=bool(request.form.get('has_itbis')),
            company_id=current_company_id()
        )
        db.session.add(product)
        db.session.commit()
        flash('Producto agregado')
        notify('Producto agregado')
        return redirect(url_for('products'))
    cat = request.args.get('cat')
    query = company_query(Product)
    if cat:
        query = query.filter_by(category=cat)
    products = query.all()
    return render_template('productos.html', products=products, units=UNITS, categories=CATEGORIES, current_cat=cat)


@app.route('/productos/importar', methods=['GET', 'POST'])
@manager_only
def products_import():
    if request.method == 'POST':
        file = request.files['file']
        rows = file.stream.read().decode('utf-8').splitlines()
        reader = csv.DictReader(rows)
        for row in reader:
            code = row.get('code')
            if not code:
                continue
            prod = company_query(Product).filter_by(code=code).first()
            if not prod:
                prod = Product(code=code, company_id=current_company_id())
                db.session.add(prod)
            prod.name = row.get('name') or prod.name
            prod.unit = row.get('unit') or prod.unit
            prod.price = _to_float(row.get('price')) or prod.price
            cat = row.get('category')
            if cat in CATEGORIES:
                prod.category = cat
            prod.has_itbis = row.get('has_itbis', '').strip().lower() in ('1', 'true', 'si', 'sí', 'yes')
            if not prod.reference:
                prod.reference = generate_reference(prod.name)
        db.session.commit()
        flash('Productos importados')
        return redirect(url_for('products'))
    return render_template('productos_importar.html')

@app.route('/productos/delete/<int:product_id>')
def delete_product(product_id):
    product = company_get(Product, product_id)
    db.session.delete(product)
    db.session.commit()
    flash('Producto eliminado')
    return redirect(url_for('products'))


@app.route('/inventario')
def inventory_report():
    wid = request.args.get('warehouse_id', type=int)
    q = request.args.get('q', '').strip()
    category = request.args.get('category', '')
    status = request.args.get('status', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)

    warehouses = company_query(Warehouse).order_by(Warehouse.name).all()
    stocks = []
    pagination = None
    movements = []
    if not wid and warehouses:
        wid = warehouses[0].id
    if wid:
        query = (
            company_query(ProductStock)
            .filter_by(warehouse_id=wid)
            .join(Product)
        )
        if q:
            like = f"%{q}%"
            query = query.filter(or_(Product.name.ilike(like), Product.code.ilike(like)))
        if category:
            query = query.filter(Product.category == category)
        if status == 'low':
            query = query.filter(ProductStock.stock > 0, ProductStock.stock <= ProductStock.min_stock)
        elif status == 'zero':
            query = query.filter(ProductStock.stock == 0)
        elif status == 'normal':
            query = query.filter(ProductStock.stock > ProductStock.min_stock)

        pagination = (
            query.order_by(Product.name)
            .paginate(page=page, per_page=per_page, error_out=False)
        )
        stocks = pagination.items
        movements = (
            company_query(InventoryMovement)
            .filter_by(warehouse_id=wid)
            .order_by(InventoryMovement.timestamp.desc())
            .limit(20)
            .all()
        )

    sales_total = (
        db.session.query(func.sum(Invoice.total))
        .filter_by(company_id=current_company_id(), warehouse_id=wid)
        .scalar()
        or 0
    )
    return render_template(
        'inventario.html',
        stocks=stocks,
        warehouses=warehouses,
        selected=wid,
        sales_total=sales_total,
        pagination=pagination,
        q=q,
        category=category,
        status=status,
        categories=CATEGORIES,
        per_page=per_page,
        movements=movements,
    )


@app.post('/inventario/<int:stock_id>/minimo')
def update_min_stock(stock_id):
    stock = company_get(ProductStock, stock_id)
    stock.min_stock = _to_int(request.form.get('min_stock'))
    db.session.commit()
    flash('Mínimo actualizado')
    return redirect(url_for('inventory_report', warehouse_id=stock.warehouse_id))


@app.route('/inventario/ajustar', methods=['GET', 'POST'])
def inventory_adjust():
    products = company_query(Product).order_by(Product.name).all()
    warehouses = company_query(Warehouse).order_by(Warehouse.name).all()
    if request.method == 'POST':
        pid = int(request.form['product_id'])
        wid = int(request.form['warehouse_id'])
        qty = _to_int(request.form['quantity'])
        mtype = request.form['movement_type']
        product = company_get(Product, pid)
        stock = company_query(ProductStock).filter_by(product_id=pid, warehouse_id=wid).first()
        if not stock:
            stock = ProductStock(product_id=pid, warehouse_id=wid, company_id=current_company_id())
            db.session.add(stock)
        # ensure numeric defaults
        if stock.stock is None:
            stock.stock = 0
        if product.stock is None:
            product.stock = 0
        if mtype == 'entrada':
            stock.stock += qty
            product.stock += qty
            mov_qty = qty
        elif mtype == 'salida':
            if stock.stock < qty:
                flash('Stock insuficiente')
                return redirect(url_for('inventory_adjust'))
            stock.stock -= qty
            product.stock -= qty
            mov_qty = qty
        else:  # ajuste
            mov_qty = abs(stock.stock - qty)
            product.stock += qty - stock.stock
            stock.stock = qty
        mov = InventoryMovement(
            product_id=product.id,
            quantity=mov_qty,
            movement_type=mtype,
            warehouse_id=wid,
            company_id=current_company_id(),
            executed_by=session.get('user_id'),
        )
        db.session.add(mov)
        db.session.commit()
        flash('Inventario actualizado')
        return redirect(url_for('inventory_report', warehouse_id=wid))
    return render_template('inventario_ajuste.html', products=products, warehouses=warehouses)


@app.route('/inventario/importar', methods=['GET', 'POST'])
def inventory_import():
    warehouses = company_query(Warehouse).order_by(Warehouse.name).all()
    if not warehouses:
        default = Warehouse(name='Principal', company_id=current_company_id())
        db.session.add(default)
        db.session.commit()
        warehouses = [default]
    if request.method == 'POST':
        wid = int(request.form['warehouse_id'])
        file = request.files.get('file')
        if not file or not file.filename.lower().endswith('.csv'):
            flash('Debe subir un archivo CSV válido')
            return render_template('inventario_importar.html', warehouses=warehouses)

        stream = StringIO(file.stream.read().decode('utf-8'))
        reader = csv.DictReader(stream)
        expected = {'code', 'stock', 'min_stock'}
        if not reader.fieldnames or not expected.issubset(set(reader.fieldnames)):
            flash('Cabeceras inválidas. Se requieren: code, stock, min_stock')
            return render_template('inventario_importar.html', warehouses=warehouses)

        errors = []
        valid_rows = []
        for idx, row in enumerate(reader, start=2):
            code = (row.get('code') or '').strip()
            if not code:
                errors.append((idx, 'Código faltante'))
                continue
            product = company_query(Product).filter_by(code=code).first()
            if not product:
                errors.append((idx, f'Producto {code} no encontrado'))
                continue
            try:
                stock_qty = int(row.get('stock'))
            except (TypeError, ValueError):
                errors.append((idx, f'Stock inválido para {code}'))
                continue
            min_val = row.get('min_stock')
            try:
                min_stock = int(min_val) if min_val not in (None, '') else None
            except ValueError:
                errors.append((idx, f'Min stock inválido para {code}'))
                continue
            valid_rows.append((product, stock_qty, min_stock))

        if errors:
            db.session.rollback()
            flash(f'Importación cancelada. {len(errors)} filas con errores.')
            return render_template('inventario_importar.html', warehouses=warehouses, errors=errors)

        for product, stock_qty, min_stock in valid_rows:
            product.stock = stock_qty
            ps = (
                company_query(ProductStock)
                .filter_by(product_id=product.id, warehouse_id=wid)
                .first()
            )
            if not ps:
                ps = ProductStock(product_id=product.id, warehouse_id=wid, company_id=current_company_id())
                db.session.add(ps)
            ps.stock = stock_qty
            if min_stock is not None:
                ps.min_stock = min_stock
                product.min_stock = min_stock
            mov = InventoryMovement(
                product_id=product.id,
                quantity=stock_qty,
                movement_type='entrada',
                reference_type='import',
                warehouse_id=wid,
                company_id=current_company_id(),
                executed_by=session.get('user_id'),
            )
            db.session.add(mov)

        db.session.commit()
        flash(f'Se importaron {len(valid_rows)} productos')
        return redirect(url_for('inventory_report', warehouse_id=wid))

    return render_template('inventario_importar.html', warehouses=warehouses)


@app.route('/inventario/transferir', methods=['GET', 'POST'])
def inventory_transfer():
    products = company_query(Product).order_by(Product.name).all()
    warehouses = company_query(Warehouse).order_by(Warehouse.name).all()
    if request.method == 'POST':
        pid = int(request.form['product_id'])
        origin = int(request.form['origin_id'])
        dest = int(request.form['dest_id'])
        qty = _to_int(request.form['quantity'])
        if origin == dest:
            flash('Seleccione almacenes distintos')
            return redirect(url_for('inventory_transfer'))
        o_stock = (
            company_query(ProductStock)
            .filter_by(product_id=pid, warehouse_id=origin)
            .first()
        )
        d_stock = (
            company_query(ProductStock)
            .filter_by(product_id=pid, warehouse_id=dest)
            .first()
        )
        if not o_stock or o_stock.stock < qty:
            flash('Stock insuficiente')
            return redirect(url_for('inventory_transfer'))
        if not d_stock:
            d_stock = ProductStock(product_id=pid, warehouse_id=dest, stock=0, company_id=current_company_id())
            db.session.add(d_stock)
        o_stock.stock -= qty
        d_stock.stock += qty
        mov_out = InventoryMovement(
            product_id=pid,
            quantity=qty,
            movement_type='salida',
            warehouse_id=origin,
            company_id=current_company_id(),
            reference_type='transfer',
            reference_id=dest,
            executed_by=session.get('user_id'),
        )
        mov_in = InventoryMovement(
            product_id=pid,
            quantity=qty,
            movement_type='entrada',
            warehouse_id=dest,
            company_id=current_company_id(),
            reference_type='transfer',
            reference_id=origin,
            executed_by=session.get('user_id'),
        )
        db.session.add_all([mov_out, mov_in])
        db.session.commit()
        flash('Transferencia realizada')
        return redirect(url_for('inventory_report', warehouse_id=dest))
    return render_template('inventario_transferir.html', products=products, warehouses=warehouses)


@app.route('/almacenes', methods=['GET', 'POST'])
@manager_only
def warehouses():
    if request.method == 'POST':
        name = request.form['name']
        address = request.form.get('address')
        w = Warehouse(name=name, address=address, company_id=current_company_id())
        db.session.add(w)
        db.session.commit()
        return redirect(url_for('warehouses'))
    ws = company_query(Warehouse).order_by(Warehouse.name).all()
    return render_template('almacenes.html', warehouses=ws)


@app.post('/almacenes/<int:w_id>/delete')
@manager_only
def delete_warehouse(w_id):
    w = company_get(Warehouse, w_id)
    db.session.delete(w)
    db.session.commit()
    return redirect(url_for('warehouses'))

@app.route('/productos/edit/<int:product_id>', methods=['GET', 'POST'])
def edit_product(product_id):
    product = company_get(Product, product_id)
    if request.method == 'POST':
        product.code = request.form['code']
        product.reference = request.form.get('reference') or generate_reference(request.form['name'])
        product.name = request.form['name']
        product.unit = request.form['unit']
        product.price = _to_float(request.form['price'])
        product.category = request.form.get('category')
        product.has_itbis = bool(request.form.get('has_itbis'))
        db.session.commit()
        flash('Producto actualizado')
        return redirect(url_for('products'))
    return render_template('producto_form.html', product=product, units=UNITS, categories=CATEGORIES)

# Quotations
@app.route('/cotizaciones')
def list_quotations():
    client_q = request.args.get('client')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    status = request.args.get('status')
    page = request.args.get('page', 1, type=int)

    cutoff = dom_now() - timedelta(days=30)
    company_query(Quotation).filter(
        Quotation.status == 'vigente', Quotation.date < cutoff
    ).update({'status': 'vencida'}, synchronize_session=False)
    db.session.commit()

    query = company_query(Quotation).join(Client)
    if client_q:
        query = query.filter(
            (Client.name.contains(client_q)) | (Client.identifier.contains(client_q))
        )
    if date_from:
        df = datetime.strptime(date_from, '%Y-%m-%d')
        query = query.filter(Quotation.date >= df)
    if date_to:
        dt = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Quotation.date < dt)
    if status:
        query = query.filter(Quotation.status == status)

    quotations = query.order_by(Quotation.date.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    return render_template(
        'cotizaciones.html',
        quotations=quotations,
        client=client_q,
        date_from=date_from,
        date_to=date_to,
        status=status,
        timedelta=timedelta,
        now=dom_now(),
    )

@app.route('/cotizaciones/nueva', methods=['GET', 'POST'])
def new_quotation():
    if request.method == 'POST':
        print('Form data:', dict(request.form))
        client_id = request.form.get('client_id')
        if not client_id:
            flash('Debe seleccionar un cliente registrado')
            return redirect(url_for('new_quotation'))
        client = company_get(Client, client_id)
        wid = request.form.get('warehouse_id')
        if not wid:
            flash('Seleccione un almacén')
            return redirect(url_for('new_quotation'))
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('product_quantity[]')
        discounts = request.form.getlist('product_discount[]')
        items = build_items(product_ids, quantities, discounts)
        if not items:
            flash('Debe agregar al menos un producto')
            return redirect(url_for('new_quotation'))
        subtotal, itbis, total = calculate_totals(items)
        payment_method = request.form.get('payment_method')
        bank = request.form.get('bank') if payment_method == 'Transferencia' else None
        quotation = Quotation(client_id=client.id, subtotal=subtotal, itbis=itbis, total=total,
                               seller=request.form.get('seller'), payment_method=payment_method,
                               bank=bank, note=request.form.get('note'),
                               warehouse_id=int(wid),
                               company_id=current_company_id())
        db.session.add(quotation)
        db.session.flush()
        for it in items:
            q_item = QuotationItem(quotation_id=quotation.id, **it)
            db.session.add(q_item)
        db.session.commit()
        flash('Cotización guardada')
        notify('Cotización guardada')
        return redirect(url_for('list_quotations'))
    clients = company_query(Client).options(
        load_only(Client.id, Client.name, Client.identifier)
    ).all()
    products = company_query(Product).options(
        load_only(Product.id, Product.code, Product.name, Product.unit, Product.price)
    ).all()
    warehouses = company_query(Warehouse).order_by(Warehouse.name).all()
    sellers = company_query(User).options(load_only(User.id, User.first_name, User.last_name)).all()
    return render_template('cotizacion.html', clients=clients, products=products, warehouses=warehouses, sellers=sellers)

@app.route('/cotizaciones/editar/<int:quotation_id>', methods=['GET', 'POST'])
def edit_quotation(quotation_id):
    quotation = company_get(Quotation, quotation_id)
    if request.method == 'POST':
        client = quotation.client
        is_final = request.form.get('client_type') == 'final'
        identifier = request.form.get('client_identifier') if not is_final else request.form.get('client_identifier') or None
        if not is_final and not identifier:
            flash('El identificador es obligatorio para comprobante fiscal')
            return redirect(url_for('edit_quotation', quotation_id=quotation.id))
        client.name = request.form['client_name']
        client.last_name = request.form.get('client_last_name') if is_final else None
        client.identifier = identifier
        client.phone = request.form.get('client_phone')
        client.email = request.form.get('client_email')
        client.street = request.form.get('client_street')
        client.sector = request.form.get('client_sector')
        client.province = request.form.get('client_province')
        client.is_final_consumer = is_final
        quotation.items.clear()
        db.session.flush()
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('product_quantity[]')
        discounts = request.form.getlist('product_discount[]')
        items = build_items(product_ids, quantities, discounts)
        subtotal, itbis, total = calculate_totals(items)
        payment_method = request.form.get('payment_method')
        bank = request.form.get('bank') if payment_method == 'Transferencia' else None
        quotation.client_id = client.id
        quotation.subtotal = subtotal
        quotation.itbis = itbis
        quotation.total = total
        quotation.seller = request.form.get('seller')
        quotation.payment_method = payment_method
        quotation.bank = bank
        quotation.note = request.form.get('note')
        for it in items:
            quotation.items.append(QuotationItem(**it))
        db.session.commit()
        flash('Cotización actualizada')
        return redirect(url_for('list_quotations'))
    products = company_query(Product).options(
        load_only(Product.id, Product.code, Product.name, Product.unit, Product.price)
    ).all()
    product_map = {p.name: p.id for p in products}
    items = []
    for it in quotation.items:
        base = it.unit_price * it.quantity
        percent = (it.discount / base * 100) if base else 0
        items.append({
            'product_id': product_map.get(it.product_name, ''),
            'quantity': it.quantity,
            'discount': percent,
            'unit': it.unit,
            'price': it.unit_price,
        })
    sellers = company_query(User).options(load_only(User.id, User.first_name, User.last_name)).all()
    return render_template(
        'cotizacion_edit.html',
        quotation=quotation,
        products=products,
        items=items,
        sellers=sellers,
    )


@app.route('/ajustes')
@manager_only
def settings():
    return redirect(url_for('settings_company'))


@app.route('/ajustes/empresa', methods=['GET', 'POST'])
@manager_only
def settings_company():
    company = CompanyInfo.query.get(current_company_id())
    if not company:
        flash('Seleccione una empresa')
        return redirect(url_for('admin_companies'))
    if request.method == 'POST':
        role = session.get('role')
        if role != 'manager':
            company.name = request.form.get('name', company.name)
            company.street = request.form.get('street', company.street)
            company.sector = request.form.get('sector', company.sector)
            company.province = request.form.get('province', company.province)
            company.phone = request.form.get('phone', company.phone)
            company.rnc = request.form.get('rnc', company.rnc)
            company.website = request.form.get('website') or None
        if request.form.get('remove_logo'):
            if company.logo:
                try:
                    os.remove(os.path.join(app.static_folder, company.logo))
                except FileNotFoundError:
                    pass
            company.logo = None
        else:
            file = request.files.get('logo')
            if file and file.filename:
                filename = secure_filename(file.filename)
                ext = os.path.splitext(filename)[1].lower()
                if ext not in {'.png', '.jpg', '.jpeg'}:
                    flash('Formato de logo inválido')
                    return redirect(url_for('settings_company'))
                file.seek(0, os.SEEK_END)
                size = file.tell()
                file.seek(0)
                if size > 1 * 1024 * 1024:
                    flash('Logo demasiado grande (máximo 1MB)')
                    return redirect(url_for('settings_company'))
                upload_dir = os.path.join(app.static_folder, 'uploads')
                os.makedirs(upload_dir, exist_ok=True)
                path = os.path.join(upload_dir, filename)
                file.save(path)
                company.logo = f'uploads/{filename}'
        old_final = company.ncf_final
        old_fiscal = company.ncf_fiscal
        new_final = _to_int(request.form.get('ncf_final'))
        new_fiscal = _to_int(request.form.get('ncf_fiscal'))
        if new_final is not None and new_final < old_final:
            flash('NCF Consumidor Final no puede ser menor que el actual')
            return redirect(url_for('settings_company'))
        if new_fiscal is not None and new_fiscal < old_fiscal:
            flash('NCF Comprobante Fiscal no puede ser menor que el actual')
            return redirect(url_for('settings_company'))
        if new_final is not None:
            company.ncf_final = new_final
        if new_fiscal is not None:
            company.ncf_fiscal = new_fiscal
        if old_final != company.ncf_final or old_fiscal != company.ncf_fiscal:
            log = NcfLog(
                company_id=company.id,
                old_final=old_final,
                old_fiscal=old_fiscal,
                new_final=company.ncf_final,
                new_fiscal=company.ncf_fiscal,
                changed_by=session.get('user_id'),
            )
            db.session.add(log)
        db.session.commit()
        flash('Ajustes guardados')
        return redirect(url_for('settings_company'))
    return render_template('ajustes_empresa.html', company=company)


@app.route('/ajustes/usuarios/agregar', methods=['GET', 'POST'])
@manager_only
def settings_add_user():
    company = CompanyInfo.query.get(current_company_id())
    if request.method == 'POST':
        if session.get('role') == 'manager':
            count = User.query.filter_by(company_id=company.id, role='company').count()
            if count >= 2:
                flash('Los managers solo pueden crear 2 usuarios')
                return redirect(url_for('settings_add_user'))
        user = User(
            username=request.form['username'],
            first_name=request.form['first_name'],
            last_name=request.form['last_name'],
            role='company',
            company_id=company.id,
        )
        user.set_password(request.form['password'])
        db.session.add(user)
        db.session.commit()
        flash('Usuario creado')
        return redirect(url_for('settings_manage_users'))
    return render_template('ajustes_usuario_form.html')


@app.route('/ajustes/usuarios', methods=['GET', 'POST'])
@manager_only
def settings_manage_users():
    company_id = current_company_id()
    if request.method == 'POST':
        uid = int(request.form['user_id'])
        user = company_get(User, uid)
        action = request.form.get('action')
        if action == 'delete':
            db.session.delete(user)
            db.session.commit()
            flash('Usuario eliminado')
            return redirect(url_for('settings_manage_users'))
        user.first_name = request.form.get('first_name', user.first_name)
        user.last_name = request.form.get('last_name', user.last_name)
        user.username = request.form.get('username', user.username)
        new_role = request.form.get('role', user.role)
        if new_role in ('company', 'manager'):
            user.role = new_role
        db.session.commit()
        flash('Usuario actualizado')
        return redirect(url_for('settings_manage_users'))
    users = (
        User.query.filter_by(company_id=company_id)
        .filter(User.id != session.get('user_id'))
        .all()
    )
    return render_template('ajustes_usuarios.html', users=users)

@app.route('/cotizaciones/<int:quotation_id>/pdf')
def quotation_pdf(quotation_id):
    quotation = company_get(Quotation, quotation_id)
    company = get_company_info()
    filename = f'cotizacion_{quotation_id}.pdf'
    pdf_path = os.path.join(app.static_folder, 'pdfs', filename)
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    valid_until = quotation.date + timedelta(days=30)
    app.logger.info("Generating quotation PDF %s", quotation_id)
    generate_pdf('Cotización', company, quotation.client, quotation.items,
                 quotation.subtotal, quotation.itbis, quotation.total,
                 seller=quotation.seller, payment_method=quotation.payment_method,
                 bank=quotation.bank, doc_number=quotation.id, note=quotation.note,
                 output_path=pdf_path,
                 date=quotation.date, valid_until=valid_until,
                 footer=("Condiciones: Esta cotización es válida por 30 días a partir de la fecha de emisión. "
                         "Los precios están sujetos a cambios sin previo aviso. "
                         "El ITBIS ha sido calculado conforme a la ley vigente."))
    return send_file(pdf_path, download_name=filename, as_attachment=True)

@app.route('/cotizaciones/<int:quotation_id>/convertir', methods=['GET', 'POST'])
def quotation_to_order(quotation_id):
    quotation = company_get(Quotation, quotation_id)
    warehouses = company_query(Warehouse).all()
    if request.method == 'GET':
        return render_template('quotation_convert.html', quotation=quotation, warehouses=warehouses)
    wid = request.form.get('warehouse_id', quotation.warehouse_id)
    if not wid:
        flash('Seleccione un almacén')
        return redirect(url_for('quotation_to_order', quotation_id=quotation_id))
    wid = int(wid)
    quotation.warehouse_id = wid
    customer_po = request.form.get('customer_po') or None
    if dom_now() > quotation.date + timedelta(days=30):
        flash('La cotización ha expirado')
        return redirect(url_for('list_quotations'))
    for item in quotation.items:
        product = company_query(Product).filter_by(code=item.code).first()
        stock = (
            company_query(ProductStock)
            .filter_by(product_id=product.id, warehouse_id=wid)
            .first()
        )
        if not stock or stock.stock < item.quantity:
            flash('Stock insuficiente para ' + item.product_name)
            return redirect(url_for('list_quotations'))
    order = Order(
        client_id=quotation.client_id,
        quotation_id=quotation.id,
        subtotal=quotation.subtotal,
        itbis=quotation.itbis,
        total=quotation.total,
        seller=quotation.seller,
        payment_method=quotation.payment_method,
        bank=quotation.bank,
        note=quotation.note,
        customer_po=customer_po,
        warehouse_id=wid,
        company_id=current_company_id(),
    )
    db.session.add(order)
    quotation.status = 'convertida'
    db.session.flush()
    for item in quotation.items:
        o_item = OrderItem(
            order_id=order.id,
            code=item.code,
            reference=item.reference,
            product_name=item.product_name,
            unit=item.unit,
            unit_price=item.unit_price,
            quantity=item.quantity,
            discount=item.discount,
            category=item.category,
            has_itbis=item.has_itbis,
            company_id=current_company_id(),
        )
        db.session.add(o_item)
        product = company_query(Product).filter_by(code=item.code).first()
        if product:
            ps = (
                company_query(ProductStock)
                .filter_by(product_id=product.id, warehouse_id=wid)
                .first()
            )
            if ps:
                ps.stock -= item.quantity
            product.stock -= item.quantity
            mov = InventoryMovement(
                product_id=product.id,
                quantity=item.quantity,
                movement_type='salida',
                reference_type='Order',
                reference_id=order.id,
                warehouse_id=wid,
                company_id=current_company_id(),
                executed_by=session.get('user_id'),
            )
            db.session.add(mov)
    db.session.commit()
    flash('Pedido creado')
    notify('Pedido creado')
    return redirect(url_for('list_orders'))

# Orders
@app.route('/pedidos')
def list_orders():
    q = request.args.get('q')
    query = company_query(Order).join(Client)
    if q:
        query = query.filter((Client.name.contains(q)) | (Client.identifier.contains(q)))
    orders = query.order_by(Order.date.desc()).all()
    return render_template('pedido.html', orders=orders, q=q)

@app.route('/pedidos/<int:order_id>/facturar')
def order_to_invoice(order_id):
    order = company_get(Order, order_id)
    company = CompanyInfo.query.get(current_company_id())
    if order.client.is_final_consumer:
        prefix, counter = "B02", "ncf_final"
    else:
        prefix, counter = "B01", "ncf_fiscal"

    # ensure NCF is unique; advance company counter until unused
    while True:
        seq = getattr(company, counter)
        ncf = f"{prefix}{seq:08d}"
        if not Invoice.query.filter_by(ncf=ncf).first():
            setattr(company, counter, seq + 1)
            break
        setattr(company, counter, seq + 1)
    invoice = Invoice(
        client_id=order.client_id,
        order_id=order.id,
        subtotal=order.subtotal,
        itbis=order.itbis,
        total=order.total,
        ncf=ncf,
        seller=order.seller,
        payment_method=order.payment_method,
        bank=order.bank,
        note=order.note,
        invoice_type=(
            'Consumidor Final' if order.client.is_final_consumer else 'Crédito Fiscal'
        ),
        status='Pendiente',
        warehouse_id=order.warehouse_id,
        company_id=current_company_id(),
    )
    db.session.add(invoice)
    db.session.flush()
    for item in order.items:
        i_item = InvoiceItem(
            invoice_id=invoice.id,
            code=item.code,
            reference=item.reference,
            product_name=item.product_name,
            unit=item.unit,
            unit_price=item.unit_price,
            quantity=item.quantity,
            discount=item.discount,
            category=item.category,
            has_itbis=item.has_itbis,
            company_id=current_company_id(),
        )
        db.session.add(i_item)
    order.status = 'Entregado'
    db.session.commit()
    flash('Factura generada')
    notify('Factura generada')
    return redirect(url_for('list_invoices'))

@app.route('/pedidos/<int:order_id>/pdf')
def order_pdf(order_id):
    order = company_get(Order, order_id)
    company = get_company_info()
    filename = f'pedido_{order_id}.pdf'
    pdf_path = os.path.join(app.static_folder, 'pdfs', filename)
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    app.logger.info("Generating order PDF %s", order_id)
    generate_pdf('Pedido', company, order.client, order.items,
                 order.subtotal, order.itbis, order.total,
                 seller=order.seller, payment_method=order.payment_method,
                 bank=order.bank, doc_number=order.id, note=order.note,
                 output_path=pdf_path,
                date=order.date,
                footer=("Este pedido será procesado tras la confirmación de pago. "
                        "Tiempo estimado de entrega: 3 a 5 días hábiles."))
    return send_file(pdf_path, download_name=filename, as_attachment=True)

# Invoices
@app.route('/facturas')
def list_invoices():
    q = request.args.get('q')
    query = company_query(Invoice).join(Client)
    if q:
        query = query.filter((Client.name.contains(q)) | (Client.identifier.contains(q)))
    invoices = query.order_by(Invoice.date.desc()).all()
    return render_template('factura.html', invoices=invoices, q=q)


@app.route('/facturas/<int:invoice_id>/pagar', methods=['POST'])
def pay_invoice(invoice_id):
    invoice = company_get(Invoice, invoice_id)
    invoice.status = 'Pagada'
    db.session.commit()
    flash('Factura marcada como pagada')
    return redirect(url_for('list_invoices'))


@app.route('/notificaciones')
def notifications_view():
    notifs = company_query(Notification).order_by(Notification.created_at.desc()).all()
    return render_template('notifications.html', notifications=notifs)


@app.post('/notificaciones/<int:nid>/leer')
def notifications_read(nid):
    notif = company_get(Notification, nid)
    notif.is_read = True
    db.session.commit()
    return redirect(url_for('notifications_view'))

@app.route('/facturas/<int:invoice_id>/pdf')
def invoice_pdf(invoice_id):
    invoice = company_get(Invoice, invoice_id)
    company = get_company_info()
    filename = f'factura_{invoice_id}.pdf'
    pdf_path = os.path.join(app.static_folder, 'pdfs', filename)
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    app.logger.info("Generating invoice PDF %s", invoice_id)
    generate_pdf('Factura', company, invoice.client, invoice.items,
                 invoice.subtotal, invoice.itbis, invoice.total,
                 ncf=invoice.ncf, seller=invoice.seller,
                 payment_method=invoice.payment_method, bank=invoice.bank,
                 purchase_order=invoice.order.customer_po if invoice.order else None,
                 doc_number=invoice.id,
                 invoice_type=invoice.invoice_type, note=invoice.note,
                 output_path=pdf_path, date=invoice.date,
                 footer=("Factura generada electrónicamente, válida sin firma ni sello. "
                         "Para reclamaciones favor comunicarse dentro de las 48 horas siguientes a la emisión. "
                         "Gracias por su preferencia."))
    return send_file(pdf_path, download_name=filename, as_attachment=True)

@app.route('/pdfs/<path:filename>')
def serve_pdf(filename):
    return send_from_directory(os.path.join(app.static_folder, 'pdfs'), filename)

def _filtered_invoice_query(fecha_inicio, fecha_fin, estado, categoria):
    """Return an invoice query filtered by the provided parameters."""
    q = company_query(Invoice)
    if fecha_inicio:
        q = q.filter(Invoice.date >= fecha_inicio)
    if fecha_fin:
        q = q.filter(Invoice.date <= fecha_fin)
    if estado:
        q = q.filter(Invoice.status == estado)
    if categoria:
        q = q.join(Invoice.items).filter(InvoiceItem.category == categoria)
    return q


@app.route('/reportes')
def reportes():
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    estado = request.args.get('estado')
    categoria = request.args.get('categoria')
    page = request.args.get('page', 1, type=int)

    start, end, estado, categoria = _parse_report_params(fecha_inicio, fecha_fin, estado, categoria)
    q = _filtered_invoice_query(start, end, estado, categoria)

    pagination = (
        q.options(
            joinedload(Invoice.client),
            load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
        )
        .order_by(Invoice.date.desc())
        .paginate(page=page, per_page=10, error_out=False)
    )
    invoices = pagination.items

    total_sales, unique_clients, invoice_count = (
        q.with_entities(
            func.coalesce(func.sum(Invoice.total), 0),
            func.count(func.distinct(Invoice.client_id)),
            func.count(Invoice.id),
        ).first()
    )

    item_query = company_query(InvoiceItem).join(Invoice)
    if start:
        item_query = item_query.filter(Invoice.date >= start)
    if end:
        item_query = item_query.filter(Invoice.date <= end)
    if estado:
        item_query = item_query.filter(Invoice.status == estado)
    if categoria:
        item_query = item_query.filter(InvoiceItem.category == categoria)

    sales_by_category = (
        item_query.with_entities(
            InvoiceItem.category,
            func.count(InvoiceItem.id),
            func.avg((InvoiceItem.unit_price * InvoiceItem.quantity) - InvoiceItem.discount),
            func.sum((InvoiceItem.unit_price * InvoiceItem.quantity) - InvoiceItem.discount),
        ).group_by(InvoiceItem.category).all()
    )

    sales_over_time = (
        q.with_entities(func.date(Invoice.date), func.sum(Invoice.total), func.count(Invoice.id))
        .group_by(func.date(Invoice.date))
        .order_by(func.date(Invoice.date))
        .all()
    )

    # retention
    client_counts = (
        q.with_entities(Invoice.client_id, func.count(Invoice.id))
        .group_by(Invoice.client_id)
        .all()
    )
    retained = len([1 for _cid, cnt in client_counts if cnt > 1])
    retention = (retained / len(client_counts)) * 100 if client_counts else 0

    # top categories last year
    last_year_start = datetime.utcnow().replace(year=datetime.utcnow().year - 1, month=1, day=1)
    top_cats = (
        company_query(InvoiceItem)
        .join(Invoice)
        .with_entities(InvoiceItem.category, func.sum((InvoiceItem.unit_price * InvoiceItem.quantity) - InvoiceItem.discount))
        .filter(Invoice.date >= last_year_start)
        .group_by(InvoiceItem.category)
        .order_by(func.sum((InvoiceItem.unit_price * InvoiceItem.quantity) - InvoiceItem.discount).desc())
        .limit(5)
        .all()
    )

    # monthly and yearly avg ticket
    today = datetime.utcnow()
    month_total, month_clients = (
        q.filter(
            func.strftime('%Y', Invoice.date) == str(today.year),
            func.strftime('%m', Invoice.date) == f"{today.month:02d}",
        )
        .with_entities(
            func.coalesce(func.sum(Invoice.total), 0),
            func.count(func.distinct(Invoice.client_id)),
        )
        .first()
    )
    avg_ticket_month = month_total / month_clients if month_clients else 0
    year_total, year_clients = (
        q.filter(func.strftime('%Y', Invoice.date) == str(today.year))
        .with_entities(
            func.coalesce(func.sum(Invoice.total), 0),
            func.count(func.distinct(Invoice.client_id)),
        )
        .first()
    )
    avg_ticket_year = year_total / year_clients if year_clients else 0

    # trend last 24 months
    trend_query = (
        q.with_entities(func.strftime('%Y-%m', Invoice.date), func.sum(Invoice.total))
        .filter(Invoice.date >= datetime(today.year - 2, today.month, 1))
        .group_by(func.strftime('%Y-%m', Invoice.date))
        .order_by(func.strftime('%Y-%m', Invoice.date))
    )
    # ensure a list of dicts is always provided even if the query returns no rows
    trend_rows = trend_query.all() if trend_query is not None else []
    trend_24 = [{'month': m, 'total': tot or 0} for m, tot in trend_rows]

    status_totals = {s: 0 for s in INVOICE_STATUSES}
    status_counts = {s: 0 for s in INVOICE_STATUSES}
    for st, amount, cnt in (
        q.with_entities(Invoice.status, func.sum(Invoice.total), func.count(Invoice.id))
        .group_by(Invoice.status)
    ):
        if st in status_totals:
            status_totals[st] = amount or 0
            status_counts[st] = cnt or 0

    payment_totals = {'Efectivo': 0, 'Transferencia': 0}
    payment_counts = {'Efectivo': 0, 'Transferencia': 0}
    for pm, amount, cnt in (
        q.with_entities(Invoice.payment_method, func.sum(Invoice.total), func.count(Invoice.id))
        .group_by(Invoice.payment_method)
    ):
        if pm in payment_totals:
            payment_totals[pm] = amount or 0
            payment_counts[pm] = cnt or 0

    current_year = datetime.utcnow().year
    monthly_totals = (
        q.with_entities(
            func.strftime('%Y', Invoice.date).label('y'),
            func.strftime('%m', Invoice.date).label('m'),
            func.sum(Invoice.total),
        )
        .filter(Invoice.date >= datetime(current_year - 1, 1, 1))
        .group_by('y', 'm')
        .all()
    )
    year_current = [0] * 12
    year_prev = [0] * 12
    for y, m, total in monthly_totals:
        if int(y) == current_year:
            year_current[int(m) - 1] = total or 0
        else:
            year_prev[int(m) - 1] = total or 0

    avg_ticket = total_sales / unique_clients if unique_clients else 0

    top_clients = (
        q.join(Client)
        .with_entities(Client.name, func.sum(Invoice.total))
        .group_by(Client.id)
        .order_by(func.sum(Invoice.total).desc())
        .limit(5)
        .all()
    )

    stats = {
        'total_sales': total_sales,
        'unique_clients': unique_clients,
        'invoices': invoice_count,
        'pending': status_totals.get('Pendiente', 0),
        'paid': status_totals.get('Pagada', 0),
        'cash': payment_totals.get('Efectivo', 0),
        'transfer': payment_totals.get('Transferencia', 0),
        'avg_ticket': avg_ticket,
        'avg_ticket_month': avg_ticket_month,
        'avg_ticket_year': avg_ticket_year,
        'retention': retention,
    }

    cat_labels = [c or 'Sin categoría' for c, *_ in sales_by_category]
    cat_totals = [s or 0 for *_1, _2, s in sales_by_category]
    cat_counts = [qtd for _cat, qtd, *_ in sales_by_category]
    date_labels = [d if isinstance(d, str) else d.strftime('%Y-%m-%d') for d, *_ in sales_over_time]
    date_totals = [t or 0 for _, t, _ in sales_over_time]
    date_counts = [cnt for *_1, cnt in sales_over_time]
    status_labels = list(status_counts.keys())
    status_values = list(status_counts.values())
    method_labels = list(payment_counts.keys())
    method_values = list(payment_counts.values())
    months = [
        'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
        'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic'
    ]

    filters = {
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'estado': estado or '',
        'categoria': categoria or '',
    }

    if request.args.get('ajax') == '1':
        return jsonify(
            {
                'stats': stats,
                'top_clients': [{'name': n, 'total': t} for n, t in top_clients],
                'cat_labels': cat_labels,
                'cat_totals': cat_totals,
                'cat_counts': cat_counts,
                'date_labels': date_labels,
                'date_totals': date_totals,
                'date_counts': date_counts,
                'status_labels': status_labels,
                'status_values': status_values,
                'method_labels': method_labels,
                'method_values': method_values,
                'months': months,
                'year_current': year_current,
                'year_prev': year_prev,
                'top_categories_year': [{'category': c or 'Sin categoría', 'total': t or 0} for c, t in top_cats],
                'trend_24': trend_24,
                'invoices': [
                    {
                        'client': i.client.name if i.client else '',
                        'date': i.date.strftime('%Y-%m-%d'),
                        'estado': i.status or '',
                        'total': i.total,
                    }
                    for i in invoices
                ],
                'pagination': {'page': pagination.page, 'pages': pagination.pages},
            }
        )

    return render_template(
        'reportes.html',
        invoices=invoices,
        pagination=pagination,
        sales_by_category=sales_by_category,
        stats=stats,
        top_clients=top_clients,
        top_categories_year=top_cats,
        trend_24=trend_24,
        cat_labels=cat_labels,
        cat_totals=cat_totals,
        cat_counts=cat_counts,
        date_labels=date_labels,
        date_totals=date_totals,
        date_counts=date_counts,
        status_labels=status_labels,
        status_values=status_values,
        method_labels=method_labels,
        method_values=method_values,
        months=months,
        year_current=year_current,
        year_prev=year_prev,
        filters=filters,
        categories=CATEGORIES,
        statuses=INVOICE_STATUSES,
    )


@app.get('/reportes/estado-cuentas')
def account_statement_clients():
    clients = company_query(Client).order_by(Client.name).all()
    return render_template('estado_cuentas.html', clients=clients)


def _invoice_balance(inv):
    return inv.total - sum(p.amount for p in inv.payments)


@app.get('/reportes/estado-cuentas/<int:client_id>')
def account_statement_detail(client_id):
    client = company_get(Client, client_id)
    invoices = (
        company_query(Invoice)
        .filter_by(client_id=client.id)
        .options(joinedload(Invoice.order), joinedload(Invoice.payments))
        .all()
    )
    rows = []
    totals = 0
    aging = {'0-30': 0, '31-60': 0, '61-90': 0, '91-120': 0, '121+': 0}
    now = datetime.utcnow()
    for inv in invoices:
        balance = _invoice_balance(inv)
        if balance <= 0:
            continue
        due = inv.date + timedelta(days=30)
        rows.append({
            'document': inv.ncf or f'FAC-{inv.id}',
            'order': inv.order.customer_po if inv.order and inv.order.customer_po else inv.order_id,
            'date': inv.date.strftime('%d/%m/%Y'),
            'due': due.strftime('%d/%m/%Y'),
            'info': inv.note or '',
            'amount': inv.total,
            'balance': balance,
        })
        totals += balance
        age = (now - inv.date).days
        if age <= 30:
            aging['0-30'] += balance
        elif age <= 60:
            aging['31-60'] += balance
        elif age <= 90:
            aging['61-90'] += balance
        elif age <= 120:
            aging['91-120'] += balance
        else:
            aging['121+'] += balance
    overdue = sum(r['balance'] for r in rows if datetime.strptime(r['due'], '%d/%m/%Y') < now)
    overdue_pct = (overdue / totals * 100) if totals else 0
    if request.args.get('pdf') == '1':
        company = {
            'name': g.company.name,
            'street': g.company.street,
            'phone': g.company.phone,
            'rnc': g.company.rnc,
            'logo': g.company.logo,
        }
        client_dict = {
            'name': client.name,
            'identifier': client.identifier,
            'street': client.street,
            'sector': client.sector,
            'province': client.province,
            'phone': client.phone,
            'email': client.email,
        }
        pdf_path = generate_account_statement_pdf(company, client_dict, rows, totals, aging, overdue_pct)
        return send_file(pdf_path, as_attachment=True, download_name=f'estado_cuenta_{client.id}.pdf')
    return render_template('estado_cuenta_detalle.html', client=client, rows=rows, total=totals, aging=aging, overdue_pct=overdue_pct)


@app.route('/reportes/export')
def export_reportes():
    role = session.get('role')
    formato = request.args.get('formato', 'csv')
    tipo = request.args.get('tipo', 'detalle')
    if role == 'contabilidad':
        if formato not in {'csv', 'xlsx'} or tipo != 'resumen':
            log_export(session.get('full_name') or session.get('username'), formato, tipo, {}, 'fail', 'permiso')
            return '', 403
    elif role not in ('admin', 'manager'):
        log_export(session.get('full_name') or session.get('username'), formato, tipo, {}, 'fail', 'permiso')
        return '', 403

    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    estado = request.args.get('estado')
    categoria = request.args.get('categoria')

    start, end, estado, categoria = _parse_report_params(fecha_inicio, fecha_fin, estado, categoria)
    q = _filtered_invoice_query(start, end, estado, categoria)
    count = q.count()
    filtros = {'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin, 'estado': estado, 'categoria': categoria}
    user = session.get('full_name') or session.get('username')

    max_rows = current_app.config.get('MAX_EXPORT_ROWS', MAX_EXPORT_ROWS)
    if count > max_rows and request.args.get('async') != '1':
        log_export(user, formato, tipo, filtros, 'fail', 'too_many_rows')
        return jsonify({'error': 'too many rows', 'suggest': 'async'}), 400

    if count > max_rows and request.args.get('async') == '1':
        entry_id = log_export(user, formato, tipo, filtros, 'queued')
        enqueue_export(
            _export_job,
            current_company_id(),
            user,
            start,
            end,
            estado,
            categoria,
            formato,
            tipo,
            entry_id,
        )
        flash('Reporte en proceso, vuelva a revisar en unos minutos')
        return jsonify({'job': entry_id})

    company = get_company_info()
    header = [
        f"Empresa: {company.get('name', '')}",
        f"Rango: {(fecha_inicio or 'Todas')} - {(fecha_fin or 'Todas')}",
        f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} por {user}",
    ]
    current_app.logger.info(
        "export user=%s company=%s formato=%s tipo=%s filtros=%s",
        user,
        current_company_id(),
        formato,
        tipo,
        filtros,
    )

    if formato == 'csv':
        if current_app.testing:
            output = StringIO()
            writer = csv.writer(output)
            for h in header:
                writer.writerow([h])
            if tipo == 'resumen':
                writer.writerow(['Categoría', 'Cantidad', 'Total'])
                summary = (
                    company_query(InvoiceItem)
                    .join(Invoice)
                    .with_entities(
                        InvoiceItem.category,
                        func.count(InvoiceItem.id),
                        func.sum(InvoiceItem.unit_price * InvoiceItem.quantity - InvoiceItem.discount),
                    )
                    .group_by(InvoiceItem.category)
                )
                if start:
                    summary = summary.filter(Invoice.date >= start)
                if end:
                    summary = summary.filter(Invoice.date <= end)
                if estado:
                    summary = summary.filter(Invoice.status == estado)
                for cat, cnt, tot in summary:
                    writer.writerow([cat or 'Sin categoría', cnt, f"{tot or 0:.2f}"])
            else:
                writer.writerow(['Cliente', 'Fecha', 'Estado', 'Total'])
                for inv in q.options(
                    joinedload(Invoice.client),
                    load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
                ):
                    writer.writerow([
                        inv.client.name if inv.client else '',
                        inv.date.strftime('%Y-%m-%d'),
                        inv.status or '',
                        f"{inv.total:.2f}",
                    ])
            mem = BytesIO()
            mem.write(output.getvalue().encode('utf-8'))
            mem.seek(0)
            log_export(user, formato, tipo, filtros, 'success')
            return send_file(mem, mimetype='text/csv', as_attachment=True, download_name='reportes.csv')
        app_obj = current_app._get_current_object()
        def generate_csv():
            with app_obj.app_context():
                sio = StringIO()
                writer = csv.writer(sio)
                for h in header:
                    writer.writerow([h])
                if tipo == 'resumen':
                    writer.writerow(['Categoría', 'Cantidad', 'Total'])
                    yield sio.getvalue(); sio.seek(0); sio.truncate(0)
                    summary = (
                        company_query(InvoiceItem)
                        .join(Invoice)
                        .with_entities(
                            InvoiceItem.category,
                            func.count(InvoiceItem.id),
                            func.sum(InvoiceItem.unit_price * InvoiceItem.quantity - InvoiceItem.discount),
                        )
                        .group_by(InvoiceItem.category)
                    )
                    if start:
                        summary = summary.filter(Invoice.date >= start)
                    if end:
                        summary = summary.filter(Invoice.date <= end)
                    if estado:
                        summary = summary.filter(Invoice.status == estado)
                    for cat, cnt, tot in summary:
                        writer.writerow([cat or 'Sin categoría', cnt, f"{tot or 0:.2f}"])
                        yield sio.getvalue(); sio.seek(0); sio.truncate(0)
                else:
                    writer.writerow(['Cliente', 'Fecha', 'Estado', 'Total'])
                    yield sio.getvalue(); sio.seek(0); sio.truncate(0)
                    stream_q = q.options(
                        joinedload(Invoice.client),
                        load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
                    ).yield_per(100)
                    for inv in stream_q:
                        writer.writerow([
                            inv.client.name if inv.client else '',
                            inv.date.strftime('%Y-%m-%d'),
                            inv.status or '',
                            f"{inv.total:.2f}",
                        ])
                        yield sio.getvalue(); sio.seek(0); sio.truncate(0)

        log_export(user, formato, tipo, filtros, 'success')
        headers = {
            'Content-Disposition': 'attachment; filename=reportes.csv'
        }
        return Response(generate_csv(), mimetype='text/csv', headers=headers)

    invoices = q.options(
        joinedload(Invoice.client),
        load_only(Invoice.client_id, Invoice.total, Invoice.date, Invoice.status),
    ).all()

    if formato == 'xlsx':
        if Workbook is None:
            mem = BytesIO()
            mem.write(b'')
            mem.seek(0)
            return send_file(mem, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='reportes.xlsx')
        wb = Workbook()
        ws = wb.active
        row = 1
        for h in header:
            ws.cell(row=row, column=1, value=h)
            row += 1
        if tipo == 'resumen':
            ws.append(['Categoría', 'Cantidad', 'Total'])
            summary = (
                company_query(InvoiceItem)
                .join(Invoice)
                .with_entities(InvoiceItem.category, func.count(InvoiceItem.id), func.sum(InvoiceItem.unit_price * InvoiceItem.quantity - InvoiceItem.discount))
                .group_by(InvoiceItem.category)
            )
            if start:
                summary = summary.filter(Invoice.date >= start)
            if end:
                summary = summary.filter(Invoice.date <= end)
            if estado:
                summary = summary.filter(Invoice.status == estado)
            for cat, cnt, tot in summary:
                ws.append([cat or 'Sin categoría', cnt, float(tot or 0)])
        else:
            ws.append(['Cliente', 'Fecha', 'Estado', 'Total'])
            for inv in invoices:
                ws.append([
                    inv.client.name if inv.client else '',
                    inv.date.strftime('%Y-%m-%d'),
                    inv.status or '',
                    float(inv.total),
                ])
        mem = BytesIO()
        wb.save(mem)
        mem.seek(0)
        log_export(user, formato, tipo, filtros, 'success')
        return send_file(
            mem,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='reportes.xlsx',
        )

    if formato == 'pdf':
        items = [
            {
                'code': inv.id,
                'reference': inv.client.name if inv.client else '',
                'product_name': inv.date.strftime('%d/%m/%Y'),
                'unit': inv.status or '',
                'unit_price': inv.total,
                'quantity': 1,
                'discount': 0,
            }
            for inv in invoices
        ]
        subtotal = sum(inv.total for inv in invoices)
        note = (
            f"Rango: {(fecha_inicio or 'Todas')} - {(fecha_fin or 'Todas')} | "
            f"Estado: {(estado or 'Todos')} | "
            f"Categoría: {(categoria or 'Todas')} | "
            f"Usuario: {user} | Facturas: {len(invoices)}"
        )
        pdf_path = generate_pdf(
            'Reporte de Facturas',
            company,
            {'name': '', 'address': '', 'phone': ''},
            items,
            subtotal,
            0,
            subtotal,
            note=note,
            output_path='reportes.pdf'
        )
        log_export(user, formato, tipo, filtros, 'success', file_path=pdf_path)
        return send_file(pdf_path, as_attachment=True, download_name='reportes.pdf')

    return redirect(url_for('reportes'))


@app.route('/reportes/inventario/export')
def export_inventory():
    role = session.get('role')
    if role not in ('admin', 'manager', 'contabilidad'):
        return '', 403
    company_id = current_company_id()
    rows = (
        db.session.query(
            Product.code,
            Product.name,
            Warehouse.name,
            ProductStock.stock,
            ProductStock.min_stock,
        )
        .join(ProductStock, Product.id == ProductStock.product_id)
        .join(Warehouse, ProductStock.warehouse_id == Warehouse.id)
        .filter(ProductStock.company_id == company_id)
        .order_by(Product.name)
        .all()
    )
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Código', 'Producto', 'Almacén', 'Stock', 'Mínimo'])
    for code, name, wh, stock, min_stock in rows:
        writer.writerow([code or '', name or '', wh or '', stock, min_stock])
    mem = BytesIO()
    mem.write(output.getvalue().encode('utf-8'))
    mem.seek(0)
    return send_file(mem, mimetype='text/csv', as_attachment=True, download_name='inventario.csv')


@app.route('/reportes/exportes')
def export_history():
    q = company_query(ExportLog).order_by(ExportLog.created_at.desc())
    usuario = request.args.get('usuario')
    formato = request.args.get('formato')
    if usuario:
        q = q.filter(ExportLog.user == usuario)
    if formato:
        q = q.filter(ExportLog.formato == formato)
    logs = q.limit(100).all()
    return render_template('export_history.html', logs=logs)


@app.route('/docs')
def docs():
    return render_template('docs.html')


@app.route('/contabilidad')
def contabilidad():
    return render_template('contabilidad.html')


@app.route('/contabilidad/catalogo')
def contab_catalogo():
    return render_template('contabilidad_catalogo.html')


@app.route('/contabilidad/entradas')
def contab_entradas():
    return render_template('contabilidad_entradas.html')


@app.route('/contabilidad/estados')
def contab_estados():
    return render_template('contabilidad_estados.html')


@app.route('/contabilidad/libro-mayor')
def contab_libro_mayor():
    return render_template('contabilidad_libro_mayor.html')


@app.route('/contabilidad/impuestos')
def contab_impuestos():
    return render_template('contabilidad_impuestos.html')


@app.route('/contabilidad/balanza')
def contab_balanza():
    return render_template('contabilidad_balanza.html')


@app.route('/contabilidad/asignacion')
def contab_asignacion():
    return render_template('contabilidad_asignacion.html')


@app.route('/contabilidad/centro-costo')
def contab_centro_costo():
    return render_template('contabilidad_centro_costo.html')


@app.route('/contabilidad/reportes')
def contab_reportes():
    return render_template('contabilidad_reportes.html')


@app.route('/contabilidad/dgii')
def contab_dgii():
    return render_template('contabilidad_dgii.html')


@app.route('/api/recommendations')
def api_recommendations():
    """Return top product recommendations based on past orders."""
    return jsonify({'products': recommend_products()})

if __name__ == '__main__':
    with app.app_context():
        ensure_admin()
    app.run(debug=True)
